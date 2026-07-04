"""Gera o template oficial de composição de rebanho por faixa etária.

Layout de célula fixa (contrato usado pelo parser universal da Tarefa 2.1):
    B1: Fazenda / Identificação
    B2: Data de referência (DD/MM/AAAA)
    B3: UF de origem (dropdown das 27 UFs)
    A5:C5  cabeçalho  -> Faixa Etária | Fêmea | Macho
    A6:C10 dados      -> 5 faixas x (Fêmea, Macho)
    B11/C11 totais    -> SOMA das colunas

Rode: python scripts/generate_template.py
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

OUTPUT = Path(__file__).resolve().parent.parent / "static" / "templates" / "modelo_composicao_rebanho.xlsx"

UFS = [
    "AC", "AL", "AP", "AM", "BA", "CE", "DF", "ES", "GO", "MA", "MT", "MS",
    "MG", "PA", "PB", "PR", "PE", "PI", "RJ", "RN", "RS", "RO", "RR", "SC",
    "SP", "SE", "TO",
]

# (rótulo, exemplo_fêmea, exemplo_macho)
FAIXAS = [
    ("00–04 meses", 12, 10),
    ("05–12 meses", 8, 9),
    ("13–24 meses", 15, 7),
    ("25–36 meses", 5, 4),
    ("Acima de 36 meses", 40, 6),
]


def build() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Composição do Rebanho"

    bold = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    center = Alignment(horizontal="center")

    # --- metadados ---
    ws["A1"] = "Fazenda / Identificação"
    ws["A2"] = "Data de referência (DD/MM/AAAA)"
    ws["A3"] = "UF de origem"
    for c in ("A1", "A2", "A3"):
        ws[c].font = bold
    ws["B1"] = "Fazenda Exemplo"
    ws["B2"] = "03/07/2026"
    ws["B3"] = "RO"

    uf_dv = DataValidation(type="list", formula1='"%s"' % ",".join(UFS), allow_blank=True)
    ws.add_data_validation(uf_dv)
    uf_dv.add(ws["B3"])

    # --- cabeçalho da tabela ---
    ws["A5"], ws["B5"], ws["C5"] = "Faixa Etária", "Fêmea", "Macho"
    for c in ("A5", "B5", "C5"):
        ws[c].font = bold
        ws[c].fill = header_fill
        ws[c].alignment = center

    # --- faixas etárias ---
    qty_dv = DataValidation(type="whole", operator="greaterThanOrEqual", formula1="0", allow_blank=True)
    ws.add_data_validation(qty_dv)
    for i, (label, ex_f, ex_m) in enumerate(FAIXAS):
        row = 6 + i
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=ex_f)
        ws.cell(row=row, column=3, value=ex_m)
        qty_dv.add(ws.cell(row=row, column=2))
        qty_dv.add(ws.cell(row=row, column=3))

    # --- totais ---
    ws["A11"] = "Total"
    ws["A11"].font = bold
    ws["B11"] = "=SUM(B6:B10)"
    ws["C11"] = "=SUM(C6:C10)"
    ws["B11"].font = ws["C11"].font = bold

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    return wb


def main() -> None:
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    build().save(OUTPUT)
    print(f"Template gerado: {OUTPUT}")


if __name__ == "__main__":
    main()
