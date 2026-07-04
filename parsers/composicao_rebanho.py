"""Leitor do template oficial de composição de rebanho (`modelo_composicao_rebanho.xlsx`).

Caminho universal de ingestão: qualquer produtor, em qualquer estado, preenche
o mesmo template padrão e o sistema lê com um único parser — sem precisar de um
parser de PDF dedicado por órgão estadual.

Layout-contrato (ver `scripts/generate_template.py`):
    B1 fazenda | B2 data de referência | B3 UF
    A5:C5 cabeçalho -> "Faixa Etária" | "Fêmea" | "Macho"
    A6:C10 dados    -> 5 faixas x (Fêmea, Macho)
"""
from __future__ import annotations

from openpyxl import load_workbook

# Linha da planilha -> prefixo da faixa no vetor/animais.
_LINHA_FAIXA = {6: "f00", 7: "f05", 8: "f13", 9: "f25", 10: "fac"}

_ORDEM = ["f00", "f05", "f13", "f25", "fac"]


def _num(valor) -> int:
    """Converte o valor de uma célula para inteiro não-negativo (0 se vazio/ inválido)."""
    if valor is None:
        return 0
    try:
        n = int(round(float(valor)))
    except (TypeError, ValueError):
        return 0
    return max(n, 0)


def ler_template(fonte) -> dict:
    """Lê uma planilha no padrão do template oficial de composição de rebanho.

    Args:
        fonte: Caminho do arquivo `.xlsx` ou objeto de arquivo (file-like).

    Returns:
        Dict no mesmo formato dos parsers de PDF, acrescido dos metadados:
        `{origem, fazenda, uf, data_saldo, total, animais, valores}`.

    Raises:
        ValueError: Se a planilha não seguir o layout do template oficial.
    """
    wb = load_workbook(fonte, data_only=True)
    ws = wb.active

    # Valida o cabeçalho-contrato antes de confiar nas posições de célula.
    if (str(ws["B5"].value).strip() != "Fêmea"
            or str(ws["C5"].value).strip() != "Macho"):
        raise ValueError(
            "Planilha fora do padrão: use o modelo_composicao_rebanho.xlsx oficial."
        )

    animais = {}
    for linha, prefixo in _LINHA_FAIXA.items():
        animais[f"{prefixo}_F"] = _num(ws.cell(row=linha, column=2).value)
        animais[f"{prefixo}_M"] = _num(ws.cell(row=linha, column=3).value)

    valores = []
    for prefixo in _ORDEM:
        valores.append(animais[f"{prefixo}_F"])
        valores.append(animais[f"{prefixo}_M"])

    fazenda = ws["B1"].value
    uf = ws["B3"].value
    data_saldo = ws["B2"].value

    return {
        "origem": "TEMPLATE",
        "fazenda": None if fazenda is None else str(fazenda).strip(),
        "uf": None if uf is None else str(uf).strip().upper(),
        "data_saldo": None if data_saldo is None else str(data_saldo).strip(),
        "total": sum(valores),
        "animais": animais,
        "valores": valores,
    }
