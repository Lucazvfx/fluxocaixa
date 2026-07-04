"""Testes do parser do template oficial de composição de rebanho."""
import pytest
from openpyxl import Workbook

from scripts.generate_template import build
from parsers.composicao_rebanho import ler_template
from services.consistencia_rebanho import analisar_consistencia


def _salvar(wb, tmp_path):
    p = tmp_path / "modelo.xlsx"
    wb.save(p)
    return p


def test_le_template_exemplo(tmp_path):
    dados = ler_template(_salvar(build(), tmp_path))
    assert dados["origem"] == "TEMPLATE"
    assert dados["uf"] == "RO"
    # Exemplos do gerador: F/M por faixa.
    assert dados["valores"] == [12, 10, 8, 9, 15, 7, 5, 4, 40, 6]
    assert dados["total"] == 116
    assert dados["animais"]["fac_F"] == 40


def test_planilha_fora_do_padrao(tmp_path):
    wb = Workbook()
    wb.active["A1"] = "planilha qualquer"
    with pytest.raises(ValueError):
        ler_template(_salvar(wb, tmp_path))


def test_valores_alimentam_analise(tmp_path):
    dados = ler_template(_salvar(build(), tmp_path))
    res = analisar_consistencia(dados["valores"])
    assert "score_consistencia" in res
    assert res["total_animais"] == dados["total"]


def test_negativos_e_vazios_viram_zero(tmp_path):
    wb = build()
    ws = wb.active
    ws["B6"] = None
    ws["C6"] = -5
    dados = ler_template(_salvar(wb, tmp_path))
    assert dados["animais"]["f00_F"] == 0
    assert dados["animais"]["f00_M"] == 0
