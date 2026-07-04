"""Garante que o template gerado mantém o layout-contrato usado pela Tarefa 2.1."""
from openpyxl import load_workbook

from scripts.generate_template import build, FAIXAS


def test_layout_contrato():
    ws = build().active
    assert ws["A5"].value == "Faixa Etária"
    assert ws["B5"].value == "Fêmea"
    assert ws["C5"].value == "Macho"
    for i, (label, _, _) in enumerate(FAIXAS):
        assert ws.cell(row=6 + i, column=1).value == label
    assert ws["A1"].value == "Fazenda / Identificação"
    assert ws["A3"].value == "UF de origem"


def test_arquivo_carrega(tmp_path):
    p = tmp_path / "modelo.xlsx"
    build().save(p)
    load_workbook(p)
