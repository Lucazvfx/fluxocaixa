"""Importação de ficha de rebanho a partir do CONSOLIDADO (xlsx/xlsm).

Formato esperado: aba CONSOLIDADO da planilha 'Classificação de Rebanho - Fichas'.
  - Col B: label feminino ('Bezerra', 'Bezerra Desmama', 'Novilha', 'Vaca')
           ou 'Fazenda' (nome da fazenda está em col C)
  - Col C: quantidade feminina (int) digitada pelo usuário, ou nome da fazenda
  - Col E: label masculino ('Bezerro', 'Bezerro Desmama', 'Garrote', 'Boi Gordo')
  - Col F: quantidade masculina (int) digitada pelo usuário

Mapeia para o array v[] de 10 posições (ordem do motor):
  [f00F, f00M, f05F, f05M, f13F, f13M, f25F, f25M, facF, facM]

  Bezerra (0-12m F)       → f00_F + f05_F (50/50)
  Bezerra Desmama (13-24m F) → f13_F
  Novilha (25-36m F)      → f25_F
  Vaca (36m+ F)           → fac_F
  Bezerro (0-12m M)       → f00_M + f05_M (50/50)
  Bezerro Desmama (13-24m M) → f13_M
  Garrote (25-36m M)      → f25_M
  Boi Gordo (36m+ M)      → fac_M

Fonte do mapeamento: MAPEAMENTO (aba da mesma planilha), estados MT/GO/MS/MA/TO/RO/PA.
"""
from __future__ import annotations
import io

_FEMALE_MAP: dict[str, tuple] = {
    'bezerra':         ('f00_F', 'f05_F'),
    'bezerra desmama': ('f13_F',),
    'novilha':         ('f25_F',),
    'vaca':            ('fac_F',),
}

_MALE_MAP: dict[str, tuple] = {
    'bezerro':         ('f00_M', 'f05_M'),
    'bezerro desmama': ('f13_M',),
    'garrote':         ('f25_M',),
    'boi gordo':       ('fac_M',),
}


def _animais_vazios() -> dict:
    return {k: 0 for k in (
        'f00_F', 'f05_F', 'f13_F', 'f25_F', 'fac_F',
        'f00_M', 'f05_M', 'f13_M', 'f25_M', 'fac_M',
    )}


def _para_valores(animais: dict) -> list:
    return [
        animais['f00_F'], animais['f00_M'],
        animais['f05_F'], animais['f05_M'],
        animais['f13_F'], animais['f13_M'],
        animais['f25_F'], animais['f25_M'],
        animais['fac_F'], animais['fac_M'],
    ]


def _add(animais: dict, keys: tuple, qtd: int) -> None:
    if len(keys) == 2:
        metade = qtd // 2
        animais[keys[0]] += metade
        animais[keys[1]] += qtd - metade
    else:
        animais[keys[0]] += qtd


def _safe_int(v) -> int:
    try:
        return max(0, int(v))
    except (TypeError, ValueError):
        return 0


_TOTAL_REBANHO_MAP: dict[str, tuple] = {
    'bezerra':         ('f00_F', 'f05_F'),
    'bezerra desmama': ('f13_F',),
    'novilha':         ('f25_F',),
    'vaca':            ('fac_F',),
    'bezerro':         ('f00_M', 'f05_M'),
    'bezerro desmama': ('f13_M',),
    'garrote':         ('f25_M',),
    'boi gordo':       ('fac_M',),
}


def parsear_ficha_excel(source) -> list[dict]:
    """Lê a aba CONSOLIDADO e devolve lista de fazendas com seus v[].

    source: caminho (str/Path), bytes ou file-like aberto em modo binário.
    Retorna: [{'fazenda': str, 'valores': list[10], 'animais': dict, 'total': int}]
    O primeiro item pode ser {'fazenda': 'Total Rebanho', ...} quando a coluna J/K
    do CONSOLIDADO tiver dados (totais agregados preenchidos pelo Excel).
    """
    import openpyxl

    if isinstance(source, (str,)) or hasattr(source, '__fspath__'):
        wb = openpyxl.load_workbook(source, keep_vba=True, data_only=True)
    elif isinstance(source, bytes):
        wb = openpyxl.load_workbook(io.BytesIO(source), keep_vba=True, data_only=True)
    else:
        wb = openpyxl.load_workbook(source, keep_vba=True, data_only=True)

    ws = wb['CONSOLIDADO'] if 'CONSOLIDADO' in wb.sheetnames else wb.active

    fazendas: list[dict] = []
    current_fazenda: str | None = None
    current_animais: dict | None = None
    total_rebanho_animais = _animais_vazios()

    for row in ws.iter_rows(values_only=True):
        # índices 0-based: A=0, B=1, C=2, D=3, E=4, F=5, J=9, K=10
        col_b = str(row[1] or '').strip() if len(row) > 1 else ''
        col_c = row[2]                      if len(row) > 2 else None
        col_e = str(row[4] or '').strip()  if len(row) > 4 else ''
        col_f = row[5]                      if len(row) > 5 else None
        col_j = str(row[9]  or '').strip() if len(row) > 9  else ''
        col_k = row[10]                     if len(row) > 10 else None

        b_lower = col_b.lower()
        e_lower = col_e.lower()
        j_lower = col_j.lower()

        if b_lower == 'fazenda':
            _flush(fazendas, current_fazenda, current_animais)
            current_fazenda = str(col_c or '').strip() or 'Sem nome'
            current_animais = _animais_vazios()
            continue

        if current_animais is None:
            continue

        if b_lower in _FEMALE_MAP and col_c is not None:
            qtd = _safe_int(col_c)
            if qtd > 0:
                _add(current_animais, _FEMALE_MAP[b_lower], qtd)

        if e_lower in _MALE_MAP and col_f is not None:
            qtd = _safe_int(col_f)
            if qtd > 0:
                _add(current_animais, _MALE_MAP[e_lower], qtd)

        # Coluna J/K: "Total Rebanho" — totais agregados pelo Excel (fórmulas)
        if j_lower in _TOTAL_REBANHO_MAP and col_k is not None:
            qtd = _safe_int(col_k)
            if qtd > 0:
                _add(total_rebanho_animais, _TOTAL_REBANHO_MAP[j_lower], qtd)

    _flush(fazendas, current_fazenda, current_animais)

    # Prefixa "Total Rebanho" (da coluna J/K) se tiver dados — aparece primeiro
    if sum(total_rebanho_animais.values()) > 0:
        valores_tr = _para_valores(total_rebanho_animais)
        fazendas.insert(0, {
            'fazenda': 'Total Rebanho',
            'animais': total_rebanho_animais,
            'valores': valores_tr,
            'total':   sum(valores_tr),
            'is_total_rebanho': True,
        })

    return fazendas


def _flush(fazendas: list, fazenda: str | None, animais: dict | None) -> None:
    if fazenda is None or animais is None:
        return
    valores = _para_valores(animais)
    if sum(valores) > 0:
        fazendas.append({
            'fazenda': fazenda,
            'animais': animais,
            'valores': valores,
            'total':   sum(valores),
        })
